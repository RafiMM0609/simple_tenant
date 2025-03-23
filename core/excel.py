from io import BytesIO
from openpyxl import load_workbook

def read_excel_column(file_content, column_name):
    with BytesIO(file_content.read()) as bytes_io:
        workbook = load_workbook(bytes_io)
        sheet = workbook.active

        column_index = None
        for cell in sheet[1]:
            if cell.value == column_name:
                column_index = cell.column
                break
        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found in Excel file.")

        column_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            column_data.append(row[column_index - 1])

    return column_data
def excel_to_list_of_dicts(file_content):
    try:
        #TODO: change to query 
        ls_execl_field=['no', 'id', 'nama', 'alamat', 'expire']
        with BytesIO(file_content.read()) as bytes_io:
            workbook = load_workbook(bytes_io)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            print('headers',headers)
            if len(ls_execl_field) == len(headers):
                validation_field = all(field in ls_execl_field for field in headers)
                if not validation_field:
                    raise ValueError(f"Error: Missing or incorrect columns. Please check and re-upload your file.")
            else:
                raise ValueError(f"Error: Missing or incorrect columns. Please check and re-upload your file.")
            data_list = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {headers[i]: row[i] for i in range(len(headers))}
                data_list.append(row_data)

            return data_list
    except ValueError as e:
        print('error :',e)
        return None

from datetime import datetime, timedelta

def excel_date_to_datetime(excel_date: int) -> datetime:
    excel_base_date = datetime(1899, 12, 30)
    return (excel_base_date + timedelta(days=excel_date)).date()